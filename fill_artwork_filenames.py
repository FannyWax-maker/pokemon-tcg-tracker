"""
fill_artwork_filenames.py
Run this in the same folder as your xlsx, with IMAGE_DIR pointing to card-images.
"""

import pandas as pd
import re
import os
from openpyxl import load_workbook

# ── CONFIG ──────────────────────────────────────────────────────────────────
XLSX_PATH   = r"C:\Users\steph\Downloads\pokemon-tcg-project\tracker-patched\public\Illustration List for App.xlsx"
IMAGE_DIR   = r"C:\Users\steph\Downloads\pokemon-tcg-project\tracker-patched\public\card-images"
OUTPUT_PATH = r"C:\Users\steph\Downloads\pokemon-tcg-project\tracker-patched\public\Illustration List for App_UPDATED.xlsx"
# ────────────────────────────────────────────────────────────────────────────

def num_to_part(num):
    if pd.isna(num): return None
    return str(num).strip().replace('/', '-').lower()

def name_to_part(name):
    if pd.isna(name): return None
    s = str(name).strip().lower()
    s = re.sub(r"['\"]", '', s)
    s = re.sub(r'\s+', '_', s)
    s = re.sub(r'[^a-z0-9_-]', '', s)
    return s

def get_set_codes(row):
    """Use only the relevant set code based on exclusivity:
    - If English set exists, use English only
    - If only Japanese, use Japanese
    - If only Chinese, use Chinese
    - If EN+JP (no CN), use both
    """
    en = row.get('English Set')
    jp = row.get('Japanese Set')
    cn = row.get('Chinese Set')
    has_en = pd.notna(en) and str(en).strip()
    has_jp = pd.notna(jp) and str(jp).strip()
    has_cn = pd.notna(cn) and str(cn).strip()

    if has_en:
        return [p.strip().lower() for p in str(en).strip().split('/') if p.strip()]
    if has_jp:
        return [p.strip().lower() for p in str(jp).strip().split('/') if p.strip()]
    if has_cn:
        return [p.strip().lower() for p in str(cn).strip().split('/') if p.strip()]
    return []

def build_candidates(row):
    num = num_to_part(row.get('Set Number'))
    if not num:
        return []

    primary = name_to_part(row.get('Pokemon'))
    if not primary:
        return []

    set_codes = get_set_codes(row)
    
    candidates = []
    suffix = f"{num}.{primary}_"
    for sc in set_codes:
        c = f"{sc}.{suffix}"
        if c not in candidates:
            candidates.append(c)
    # No-prefix fallback
    c = f".{suffix}"
    if c not in candidates:
        candidates.append(c)

    return candidates

# Load image stems
print(f"Scanning: {IMAGE_DIR}")
image_stems = {}
for f in os.listdir(IMAGE_DIR):
    stem = os.path.splitext(f)[0]
    image_stems[stem] = f
print(f"Found {len(image_stems)} images\n")

# Load spreadsheet
df = pd.read_excel(XLSX_PATH, header=None)
df.columns = df.iloc[0]
df = df.iloc[1:].reset_index(drop=True)

matched = 0
unmatched = 0
already_filled = 0
results = []

for i, row in df.iterrows():
    existing = row.get('Artwork Filename')
    if pd.notna(existing) and str(existing).strip():
        results.append(existing)
        already_filled += 1
        continue

    if pd.isna(row.get('Owned')):
        results.append(None)
        continue

    candidates = build_candidates(row)
    found = None
    for cand in candidates:
        if cand in image_stems:
            found = image_stems[cand]
            break

    results.append(found)
    if found:
        matched += 1
    else:
        unmatched += 1
        print(f"  NO MATCH: {row.get('Pokemon')} | {row.get('Set Number')} | tried: {candidates[:4]}")

print(f"\n✅ Matched:        {matched}")
print(f"⏭️  Already filled: {already_filled}")
print(f"❌ Unmatched:      {unmatched}")

# Write back
wb = load_workbook(XLSX_PATH)
ws = wb.active
for i, val in enumerate(results):
    ws.cell(row=i+2, column=13, value=val)

wb.save(OUTPUT_PATH)
print(f"\n💾 Saved to {OUTPUT_PATH}")
